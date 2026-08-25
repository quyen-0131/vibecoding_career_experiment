"""Generate a compact, reviewable O*NET index for the local prototype.

Usage:
  python scripts/generate-onet-taxonomy.py <directory-with-onet-xlsx-files>

The source workbooks stay outside the UI bundle. Only the small generated JSON
file is imported by the app. O*NET is evidence about occupations; it is not
treated as proof that a resume sentence demonstrates a skill.
"""

from __future__ import annotations

import json
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "data" / "raw" / "onet"
OUTPUT = ROOT / "data" / "generated" / "onet-occupations.json"


def rows(filename: str):
    workbook = load_workbook(SOURCE / filename, read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    headers = [str(value) for value in next(iterator)]
    for values in iterator:
        yield dict(zip(headers, values))


def importance_rows(filename: str):
    grouped: dict[str, list[dict]] = defaultdict(list)
    for row in rows(filename):
        if row.get("Scale Name") != "Importance" or row.get("Data Value") is None:
            continue
        grouped[str(row["O*NET-SOC Code"])].append(
            {
                "id": str(row["Element ID"]),
                "label": str(row["Element Name"]),
                "importance": round(float(row["Data Value"]), 2),
            }
        )
    return {
        code: sorted(items, key=lambda item: (-item["importance"], item["label"]))[:10]
        for code, items in grouped.items()
    }


def task_rows():
    grouped: dict[str, list[dict]] = defaultdict(list)
    for row in rows("Task Statements.xlsx"):
        grouped[str(row["O*NET-SOC Code"])].append(
            {
                "id": str(row["Task ID"]),
                "label": str(row["Task"]),
                "type": str(row.get("Task Type") or ""),
            }
        )
    return {
        code: sorted(items, key=lambda item: (item["type"].lower() != "core", item["id"]))[:10]
        for code, items in grouped.items()
    }


def main():
    required = [
        "Occupation Data.xlsx",
        "Task Statements.xlsx",
        "Work Activities.xlsx",
        "Essential Skills.xlsx",
        "Transferable Skills.xlsx",
    ]
    missing = [name for name in required if not (SOURCE / name).exists()]
    if missing:
        raise SystemExit("Missing O*NET workbooks: " + ", ".join(missing))

    tasks = task_rows()
    work_activities = importance_rows("Work Activities.xlsx")
    essential_skills = importance_rows("Essential Skills.xlsx")
    transferable_skills = importance_rows("Transferable Skills.xlsx")

    occupations = []
    for row in rows("Occupation Data.xlsx"):
        code = str(row["O*NET-SOC Code"])
        occupations.append(
            {
                "id": code,
                "title": str(row["Title"]),
                "description": str(row["Description"]),
                "tasks": tasks.get(code, []),
                "workActivities": work_activities.get(code, []),
                "essentialSkills": essential_skills.get(code, []),
                "transferableSkills": transferable_skills.get(code, []),
            }
        )

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(
        json.dumps(
            {
                "source": "O*NET 30.3",
                "generatedFrom": required,
                "occupations": occupations,
            },
            ensure_ascii=False,
            separators=(",", ":"),
        ),
        encoding="utf-8",
    )
    print(f"Generated {len(occupations)} occupations at {OUTPUT}")


if __name__ == "__main__":
    main()
